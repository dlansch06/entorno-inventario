import io
import json
import os
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from django.conf import settings
from django.contrib import admin
from unfold.admin import ModelAdmin
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.utils.html import format_html  

from .models import Actividad, Asistencia, Comunicado, Designacion, Equipo, Estudiante, Nota, Perfil


@admin.action(description='Descargar Reporte Excel (Oficial)')
def exportar_excel_pro(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    header_font = Font(name='Arial', bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4B2C20", end_color="4B2C20", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    logo_path = os.path.join(settings.BASE_DIR, 'static/img/logo.jpg') 
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width, img.height = 80, 80
        ws.add_image(img, 'A1') 

    ws.merge_cells('B2:F2')
    ws['B2'] = "I.E. JUANA CERVANTES - REPORTE DE INVENTARIO"
    ws['B2'].font = Font(size=14, bold=True)
    ws['B2'].alignment = center_align

    headers = ['SERIE', 'CATEGORÍA', 'MARCA', 'MODELO', 'ESTADO', 'INGRESO COLEGIO']
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for row_num, obj in enumerate(queryset, 7):
        data = [
            obj.serie, obj.get_categoria_display(), obj.marca, obj.modelo,
            obj.estado, obj.fecha_ingreso_colegio.strftime('%d/%m/%Y') if obj.fecha_ingreso_colegio else ""
        ]
        for col_num, cell_value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.border = border

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.xlsx"'
    wb.save(response)
    return response


@admin.register(Equipo)
class EquipoAdmin(ModelAdmin):
    list_display = ('serie', 'marca', 'modelo', 'color_estado', 'aula_actual', 'fecha_ingreso_colegio')
    list_filter = ('estado', 'categoria', 'fecha_ingreso_colegio')
    search_fields = ('serie', 'marca', 'modelo')
    actions = [exportar_excel_pro]

    def color_estado(self, obj):
        colores = {
            'DISPONIBLE': 'green',
            'EN_USO': 'blue',
            'MANTENIMIENTO': 'orange',
            'BAJA': 'red',
        }
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            colores.get(obj.estado, 'black'),
            obj.get_estado_display()
        )
    color_estado.short_description = 'Estado Actual'


@admin.action(description='Aprobar usuarios seleccionados')
def aprobar_usuarios(modeladmin, request, queryset):
    queryset.update(esta_aprobado=True, fecha_aprobacion=timezone.now())


@admin.register(Perfil)
class PerfilAdmin(ModelAdmin):
    list_display = ('user', 'rol', 'esta_aprobado', 'grado_asignado', 'seccion_asignada', 'verificacion_sistema')
    list_filter = ('rol', 'esta_aprobado', 'grado_asignado')
    list_editable = ('esta_aprobado',)

    def verificacion_sistema(self, obj):
        if obj.rol != 'PADRE':
            return "-"
        if not obj.dni_hija:
            return format_html('<span style="color: red; font-weight: bold;">{}</span>', "Sin datos")
        
        estudiante_existe = Estudiante.objects.filter(dni=obj.dni_hija).exists()
        if estudiante_existe:
            return format_html('<span style="color: green; font-weight: bold;">{}</span>', "Aprobado en Colegio")
        else:
            return format_html('<span style="color: red; font-weight: bold;">{}</span>', "DNI No Encontrado")
            
    verificacion_sistema.short_description = 'Validacion Automatica'

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        if not request.user.is_authenticated or not hasattr(request.user, 'perfil'):
            return qs.none()
        perfil_usuario = request.user.perfil
        if perfil_usuario.rol == 'DIRECTOR':
            return qs.filter(rol__in=['PADRE', 'DOCENTE'])
        if perfil_usuario.rol == 'SECRETARIA':
            return qs.filter(rol__in=['DOCENTE', 'PADRE'])
        return qs.none()


@admin.register(Estudiante)
class EstudianteAdmin(ModelAdmin):
    list_display = ('apellidos', 'nombres', 'dni', 'nivel', 'grado', 'seccion', 'padre')
    list_filter = ('nivel', 'grado', 'seccion')
    search_fields = ('dni', 'apellidos')
    ordering = ('nivel', 'grado', 'seccion', 'apellidos')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        if not request.user.is_authenticated or not hasattr(request.user, 'perfil'):
            return qs.none()
        perfil = request.user.perfil
        if perfil.rol in ['DIRECTOR', 'SECRETARIA', 'ADMIN']:
            return qs
        if perfil.rol == 'DOCENTE':
            return qs.filter(
                nivel=perfil.nivel_asignado,
                grado=perfil.grado_asignado,
                seccion=perfil.seccion_asignada
            )
        return qs.none()


@admin.register(Designacion)
class DesignacionAdmin(ModelAdmin):
    list_display = ('equipo', 'persona_responsable', 'fecha_entrega', 'fecha_devolucion_real')
    list_filter = ('fecha_entrega', 'fecha_devolucion_real')


@admin.register(Nota)
class NotaAdmin(ModelAdmin):
    list_display = ('estudiante', 'curso', 'b1', 'b2', 'b3', 'b4')
    list_editable = ('b1', 'b2', 'b3', 'b4')
    list_filter = ('curso', 'estudiante__nivel', 'estudiante__grado', 'estudiante__seccion')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        if not request.user.is_authenticated or not hasattr(request.user, 'perfil'):
            return qs.none()
        perfil = request.user.perfil
        if perfil.rol in ['ADMIN', 'SECRETARIA', 'DIRECTOR']:
            return qs
        if perfil.rol == 'DOCENTE':
            return qs.filter(
                estudiante__nivel=perfil.nivel_asignado,
                estudiante__grado=perfil.grado_asignado,
                estudiante__seccion=perfil.seccion_asignada
            )
        return qs.none()


@admin.register(Asistencia)
class AsistenciaAdmin(ModelAdmin):
    list_display = ('estudiante', 'fecha', 'estado')
    list_filter = ('fecha', 'estado', 'estudiante__grado')
    list_editable = ('estado',)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        if not request.user.is_authenticated or not hasattr(request.user, 'perfil'):
            return qs.none()
        perfil = request.user.perfil
        if perfil.rol in ['DIRECTOR', 'SECRETARIA', 'ADMIN']:
            return qs
        if perfil.rol == 'DOCENTE':
            return qs.filter(
                estudiante__nivel=perfil.nivel_asignado,
                estudiante__grado=perfil.grado_asignado,
                estudiante__seccion=perfil.seccion_asignada
            )
        return qs.none()


@admin.register(Comunicado)
class ComunicadoAdmin(ModelAdmin):
    list_display = ('titulo', 'autor', 'es_general', 'fecha_publicacion')
    list_filter = ('es_general', 'fecha_publicacion')
    search_fields = ('titulo', 'contenido')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        if not request.user.is_authenticated or not hasattr(request.user, 'perfil'):
            return qs.none()
        perfil = request.user.perfil
        if perfil.rol in ['DIRECTOR', 'SECRETARIA', 'ADMIN']:
            return qs
        if perfil.rol == 'DOCENTE':
            return qs.filter(autor=request.user, es_general=False)
        return qs.none()

    def save_model(self, request, obj, form, change):
        if not request.user.is_authenticated or not hasattr(request.user, 'perfil'):
            super().save_model(request, obj, form, change)
            return
        if request.user.perfil.rol == 'DOCENTE':
            obj.es_general = False
            obj.autor = request.user
        elif not change: 
            obj.autor = request.user
        super().save_model(request, obj, form, change)


@admin.register(Actividad)
class ActividadAdmin(ModelAdmin):
    list_display = ('titulo', 'tipo', 'fecha_actividad')
    list_filter = ('tipo', 'fecha_actividad')
    search_fields = ('titulo', 'descripcion')

    def has_module_permission(self, request):
        if request.user.is_superuser:
            return True
        if not request.user.is_authenticated or not hasattr(request.user, 'perfil'):
            return False
        return request.user.perfil.rol in ['DIRECTOR', 'SECRETARIA', 'ADMIN']