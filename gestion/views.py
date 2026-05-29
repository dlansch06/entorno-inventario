import io
import json
from datetime import datetime, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login as auth_login, authenticate, logout as auth_logout
from django.contrib.auth.forms import AuthenticationForm
from django.db.models import Q, Count
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import ensure_csrf_cookie
from django.urls import reverse
from .models import Equipo, Estudiante, Actividad, Comunicado, Asistencia, Nota, Perfil
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


FAKE_NOTAS = [
    {
        'curso': 'Matemática',
        'b1': 15.0,
        'b2': 16.5,
        'b3': 14.5,
        'b4': 17.0,
        'promedio': 15.75
    },
    {
        'curso': 'Comunicación',
        'b1': 14.0,
        'b2': 15.0,
        'b3': 13.5,
        'b4': 16.0,
        'promedio': 14.625
    },
    {
        'curso': 'Historia',
        'b1': 17.0,
        'b2': 18.0,
        'b3': 16.5,
        'b4': 19.0,
        'promedio': 17.625
    },
]

FAKE_ASISTENCIAS = [
    {'fecha': timezone.now().date(), 'estado': 'Presente'},
    {'fecha': timezone.now().date() - timedelta(days=1), 'estado': 'Tardanza'},
    {'fecha': timezone.now().date() - timedelta(days=2), 'estado': 'Ausente'},
]

@ensure_csrf_cookie
def login_view(request):
    if request.user.is_authenticated:
        if request.user.is_staff or request.user.is_superuser:
            return redirect('/admin/')
        return redirect('dashboard')

    if request.method == 'POST':
        action_type = request.POST.get('action_type', 'login_habitual')

        if action_type == 'login_habitual':
            form = AuthenticationForm(data=request.POST)
            if form.is_valid():
                user = form.get_user()
                auth_login(request, user)
                if user.is_staff or user.is_superuser:
                    return redirect('/admin/')
                if 'datos_reales' in request.session:
                    del request.session['datos_reales']
                return redirect('dashboard')
            else:
                messages.error(request, "Usuario o contraseña incorrectos.")
                return render(request, 'login.html')

        elif action_type == 'registro_activacion':
            dni_padre = request.POST.get('dni_padre', '').strip()
            nombre_completo = request.POST.get('nombre_completo', '').strip()
            email_padre = request.POST.get('email_padre', '').strip()
            dni_hija = request.POST.get('dni_hija', '').strip()
            nueva_password = request.POST.get('nueva_password', '').strip()
            next_url = request.POST.get('next', '')

            if not all([dni_padre, nombre_completo, dni_hija, nueva_password]):
                messages.error(request, "Todos los campos son obligatorios.")
                return render(request, 'login.html')

            partes = nombre_completo.split()
            if len(partes) >= 2:
                first_name = partes[0]
                last_name = ' '.join(partes[1:])
            else:
                first_name = nombre_completo
                last_name = ''
            apellido_padre = last_name if last_name else first_name

            try:
                estudiante = Estudiante.objects.get(dni=dni_hija)
            except Estudiante.DoesNotExist:
                messages.error(request, "El DNI de la estudiante no pertenece a ninguna alumna matriculada.")
                return render(request, 'login.html')

            apellidos_estudiante = estudiante.apellidos.lower()
            if apellido_padre.lower() not in apellidos_estudiante:
                messages.error(request, "El apellido del padre no coincide con los apellidos de la estudiante.")
                return render(request, 'login.html')

            user, created = User.objects.get_or_create(
                username=dni_padre,
                defaults={
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email_padre
                }
            )
            user.set_password(nueva_password)
            if not created:
                user.first_name = first_name
                user.last_name = last_name
                user.email = email_padre
            user.save()

            perfil, _ = Perfil.objects.get_or_create(user=user)
            perfil.rol = 'PADRE'
            perfil.dni = dni_padre
            perfil.dni_hija = dni_hija
            perfil.nivel_asignado = estudiante.nivel
            perfil.grado_asignado = estudiante.grado
            perfil.seccion_asignada = estudiante.seccion
            perfil.esta_aprobado = True
            perfil.fecha_aprobacion = timezone.now()
            perfil.save()

            estudiante.padre = user
            estudiante.save()

            user_autenticado = authenticate(request, username=dni_padre, password=nueva_password)
            if user_autenticado:
                auth_login(request, user_autenticado)
                messages.success(request, "¡Registro exitoso! Bienvenido al sistema.")
                if next_url:
                    return redirect(next_url)
                return redirect('dashboard')  # ← Cambiado de 'notas' a 'dashboard'
            else:
                messages.error(request, "Error al iniciar sesión automáticamente. Intenta manualmente.")
                return redirect('login')

        return render(request, 'login.html')

    return render(request, 'login.html')


def registro_padre_form(request):
    """Muestra el formulario de registro de padres."""
    next_url = request.GET.get('next', '')
    return redirect(f"{reverse('login')}?modo=registro&next={next_url}")


def invitado_view(request):
    request.session['es_invitado'] = True
    return redirect('dashboard')


def logout_view(request):
    auth_logout(request)
    request.session.flush()
    return redirect('login')


def dashboard_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')
    return render(request, 'dashboard.html')


def acerca_de_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')
    return render(request, 'info.html')


def notas_view(request):
    if request.GET.get('exportar') in ('pdf', 'excel') and request.GET.get('estudiante_id'):
        estudiante_id = request.GET.get('estudiante_id')
        bimestre = request.GET.get('bimestre')
        if bimestre == 'todos':
            bimestre = None
        if request.GET.get('exportar') == 'pdf':
            return exportar_libreta_pdf(request, estudiante_id, bimestre)
        else:
            return exportar_libreta_excel(request, estudiante_id, bimestre)

    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect(f"{reverse('registro_padre')}?next={request.path}")

    notas_a_mostrar = FAKE_NOTAS.copy()
    estudiante = None
    error_msg = None
    es_invitado = not request.user.is_authenticated
    perfil = None
    son_datos_falsos = True
    promedios_bimestres = {}

    if request.user.is_authenticated:
        perfil = getattr(request.user, 'perfil', None)
        if not perfil:
            error_msg = "No existe perfil asociado."
        elif perfil.rol != 'PADRE':
            error_msg = "Este panel es exclusivo para padres."

    if request.method == 'POST':
        dni_manual = request.POST.get('documento')
        nombre_manual = request.POST.get('nombre')

        if dni_manual and nombre_manual:
            if es_invitado or not request.user.is_authenticated:
                error_msg = "Debes registrarte como padre para consultar calificaciones reales. Crea una cuenta e inicia sesión."
                notas_a_mostrar = FAKE_NOTAS
                estudiante = None
                son_datos_falsos = True
            else:
                try:
                    estudiante_encontrado = Estudiante.objects.get(dni=dni_manual)
                    nombre_completo = f"{estudiante_encontrado.nombres} {estudiante_encontrado.apellidos}".lower()
                    if nombre_manual.lower() in nombre_completo:
                        estudiante = estudiante_encontrado
                        notas_reales = Nota.objects.filter(estudiante=estudiante)
                        if notas_reales.exists():
                            notas_a_mostrar = [
                                {
                                    'curso': n.curso,
                                    'b1': float(n.b1),
                                    'b2': float(n.b2),
                                    'b3': float(n.b3),
                                    'b4': float(n.b4),
                                    'promedio': float(n.promedio)
                                }
                                for n in notas_reales
                            ]
                            son_datos_falsos = False
                            request.session['datos_reales'] = True
                        else:
                            notas_a_mostrar = []
                            son_datos_falsos = False
                            error_msg = "La estudiante no tiene notas registradas todavía."
                    else:
                        estudiante = None
                        notas_a_mostrar = FAKE_NOTAS
                        son_datos_falsos = True
                        error_msg = "El nombre no coincide con el número de DNI de la estudiante."
                except Estudiante.DoesNotExist:
                    estudiante = None
                    notas_a_mostrar = FAKE_NOTAS
                    son_datos_falsos = True
                    error_msg = "No se encontró ninguna estudiante con el DNI ingresado."

    if notas_a_mostrar and not son_datos_falsos:
        cantidad = len(notas_a_mostrar)
        if cantidad > 0:
            b1_sum = sum(n['b1'] for n in notas_a_mostrar)
            b2_sum = sum(n['b2'] for n in notas_a_mostrar)
            b3_sum = sum(n['b3'] for n in notas_a_mostrar)
            b4_sum = sum(n['b4'] for n in notas_a_mostrar)
            promedios_bimestres = {
                'b1': round(b1_sum / cantidad, 1),
                'b2': round(b2_sum / cantidad, 1),
                'b3': round(b3_sum / cantidad, 1),
                'b4': round(b4_sum / cantidad, 1),
            }
    elif notas_a_mostrar and son_datos_falsos:
        cantidad = len(notas_a_mostrar)
        if cantidad > 0:
            b1_sum = sum(n['b1'] for n in notas_a_mostrar)
            b2_sum = sum(n['b2'] for n in notas_a_mostrar)
            b3_sum = sum(n['b3'] for n in notas_a_mostrar)
            b4_sum = sum(n['b4'] for n in notas_a_mostrar)
            promedios_bimestres = {
                'b1': round(b1_sum / cantidad, 1),
                'b2': round(b2_sum / cantidad, 1),
                'b3': round(b3_sum / cantidad, 1),
                'b4': round(b4_sum / cantidad, 1),
            }
    promedio_base = 0
    if notas_a_mostrar:
        suma = sum(n['promedio'] for n in notas_a_mostrar)
        promedio_base = round(suma / len(notas_a_mostrar), 1)

    return render(
        request,
        'academico/notas.html',
        {
            'estudiante': estudiante,
            'perfil': perfil,
            'notas': notas_a_mostrar,
            'error': error_msg,
            'es_invitado': es_invitado,
            'promedio_base': promedio_base,
            'son_datos_falsos': son_datos_falsos,
            'promedios_bimestres': promedios_bimestres,
        }
    )


def asistencias_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect(f"{reverse('registro_padre')}?next={request.path}")

    es_invitado = not request.user.is_authenticated
    asistencias_a_mostrar = FAKE_ASISTENCIAS.copy()
    estudiante = None
    error_msg = None
    son_datos_falsos = True

    stats = {
        'presentes': 142,
        'tardanzas': 5,
        'faltas': 3,
        'porcentaje': 96.6
    }

    if request.method == 'POST':
        dni_manual = request.POST.get('documento')
        nombre_manual = request.POST.get('nombre')

        if dni_manual and nombre_manual:
            if es_invitado or not request.user.is_authenticated:
                error_msg = "Debes registrarte como padre para consultar asistencias reales."
                asistencias_a_mostrar = FAKE_ASISTENCIAS
                son_datos_falsos = True
            else:
                try:
                    estudiante_encontrado = Estudiante.objects.get(dni=dni_manual)
                    nombre_completo = f"{estudiante_encontrado.nombres} {estudiante_encontrado.apellidos}".lower()
                    if nombre_manual.lower() in nombre_completo:
                        estudiante = estudiante_encontrado
                        asistencias_reales = Asistencia.objects.filter(estudiante=estudiante).order_by('-fecha')
                        if asistencias_reales.exists():
                            asistencias_a_mostrar = [
                                {'fecha': a.fecha.strftime('%d/%m/%Y'), 'estado': a.get_estado_display()}
                                for a in asistencias_reales
                            ]
                            son_datos_falsos = False
                            presentes = asistencias_reales.filter(estado='PRESENTE').count()
                            tardanzas = asistencias_reales.filter(estado='TARDANZA').count()
                            faltas = asistencias_reales.filter(estado='AUSENTE').count()
                            total_dias = presentes + tardanzas + faltas
                            porcentaje = 100.0 if total_dias == 0 else round(((presentes + tardanzas) / total_dias) * 100, 1)
                            stats = {
                                'presentes': presentes,
                                'tardanzas': tardanzas,
                                'faltas': faltas,
                                'porcentaje': porcentaje
                            }
                        else:
                            asistencias_a_mostrar = []
                            son_datos_falsos = False
                            error_msg = "No hay registros de asistencia para esta estudiante."
                            stats = {'presentes': 0, 'tardanzas': 0, 'faltas': 0, 'porcentaje': 0.0}
                    else:
                        asistencias_a_mostrar = FAKE_ASISTENCIAS
                        son_datos_falsos = True
                        error_msg = "El nombre no coincide con el DNI."
                except Estudiante.DoesNotExist:
                    asistencias_a_mostrar = FAKE_ASISTENCIAS
                    son_datos_falsos = True
                    error_msg = "DNI no encontrado."

    return render(request, 'academico/asistencias.html', {
        'asistencias': asistencias_a_mostrar,
        'es_invitado': es_invitado,
        'error': error_msg,
        'son_datos_falsos': son_datos_falsos,
        'estudiante': estudiante,
        'stats': stats,
    })


def comunicados_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect(f"{reverse('registro_padre')}?next={request.path}")

    es_invitado = request.session.get('es_invitado', False)
    esta_autenticado = request.user.is_authenticated

    if esta_autenticado and hasattr(request.user, 'perfil') and request.user.perfil.rol == 'PADRE' and request.user.perfil.esta_aprobado:
        try:
            estudiante = Estudiante.objects.get(dni=request.user.perfil.dni_hija)
            comunicados = Comunicado.objects.filter(
                Q(es_general=True) |
                Q(tipo='AULA', grado=estudiante.grado, seccion=estudiante.seccion)
            ).order_by('-fecha_publicacion')
        except Estudiante.DoesNotExist:
            comunicados = Comunicado.objects.filter(es_general=True).order_by('-fecha_publicacion')
        user_aprobado = True
    elif esta_autenticado:
        # Personal (director, secretaria, docente) ve todos
        comunicados = Comunicado.objects.all().order_by('-fecha_publicacion')
        user_aprobado = True
    else:
        comunicados = Comunicado.objects.filter(es_general=True).order_by('-fecha_publicacion')
        user_aprobado = False

    return render(request, 'academico/comunicados.html', {
        'comunicados': comunicados,
        'user_aprobado': user_aprobado,
        'es_invitado': es_invitado
    })


def actividades_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')
    hoy = timezone.now().date()
    hace_15_dias = hoy - timedelta(days=15)
    proximos_eventos = Actividad.objects.filter(
        fecha_actividad__date__gte=hoy
    ).order_by('fecha_actividad')
    eventos_recientes = Actividad.objects.filter(
        fecha_actividad__date__gte=hace_15_dias,
        fecha_actividad__date__lt=hoy
    ).order_by('-fecha_actividad')
    return render(request, 'academico/actividades.html', {
        'proximos_eventos': proximos_eventos,
        'eventos_recientes': eventos_recientes
    })


def panel_invitado_calendario(request):
    actividades = Actividad.objects.filter(
        fecha_actividad__gte=timezone.now()
    ).order_by('fecha_actividad')
    return render(request, 'academico/actividades.html', {'actividades': actividades})


def nivel_academico(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')
    return render(request, 'academico/nivel_academico.html', {'nivel': 'SEC', 'nombre_nivel': 'Secundaria'})


def inventario_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')

    equipos = Equipo.objects.filter(eliminado=False)
    
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    categoria = request.GET.get('categoria', '')

    if q:
        equipos = equipos.filter(
            Q(serie__icontains=q) | Q(marca__icontains=q) | Q(modelo__icontains=q)
        )
    if estado:
        equipos = equipos.filter(estado=estado)
    if categoria:
        equipos = equipos.filter(categoria=categoria)

    activos_totales = Equipo.objects.filter(eliminado=False)
    
    total_count = activos_totales.count()
    count_disponible = activos_totales.filter(estado='DISPONIBLE').count()
    count_en_uso = activos_totales.filter(estado='EN_USO').count()
    count_mantenimiento = activos_totales.filter(estado='MANTENIMIENTO').count()
    count_baja = activos_totales.filter(estado='BAJA').count()

    labels_grafico = [cat[1] for cat in Equipo.CATEGORIAS]
    
    conteo_por_cat = {cat[0]: 0 for cat in Equipo.CATEGORIAS}
    
    agrupacion_bd = activos_totales.values('categoria').annotate(total=Count('id'))
    for item in agrupacion_bd:
        cat_key = item['categoria']
        if cat_key in conteo_por_cat:
            conteo_por_cat[cat_key] = item['total']
            
    data_grafico = [conteo_por_cat[cat[0]] for cat in Equipo.CATEGORIAS]

    return render(request, 'inventario.html', {
        'equipos': equipos,  
        'q': q,
        'estado_actual': estado,
        'categoria_actual': categoria,
        'categorias': Equipo.CATEGORIAS,
        'estados': Equipo.ESTADOS,
        
        'total_count': total_count,
        'count_disponible': count_disponible,
        'count_en_uso': count_en_uso,
        'count_mantenimiento': count_mantenimiento,
        'count_baja': count_baja,
        
        'chart_labels': json.dumps(labels_grafico),
        'chart_data': json.dumps(data_grafico),
    })

def reportes_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')

    equipos = Equipo.objects.filter(eliminado=False)
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')

    if q:
        equipos = equipos.filter(
            Q(serie__icontains=q) |
            Q(marca__icontains=q) |
            Q(modelo__icontains=q)
        )
    if estado:
        equipos = equipos.filter(estado=estado)

    export_type = request.GET.get('export')
    if export_type == 'excel':
        return exportar_excel(equipos)
    elif export_type == 'pdf':
        return exportar_pdf(equipos)

    return render(request, 'reportes.html', {
        'equipos': equipos,
        'q': q,
        'estado_actual': estado
    })


def detalle_equipo_view(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk, eliminado=False)
    historial = equipo.historial.all().order_by('-fecha_entrega')
    ultima_asignacion = historial.filter(fecha_devolucion_real__isnull=True).first()
    context = {
        'equipo': equipo,
        'historial': historial,
        'ultima_asignacion': ultima_asignacion,
    }
    return render(request, 'detalle_equipo.html', context)


@login_required
def mover_a_papelera(request, pk):
    if not request.user.is_staff:
        messages.error(request, "No tienes permiso para eliminar equipos.")
        return redirect('inventario')
    equipo = get_object_or_404(Equipo, pk=pk)
    equipo.eliminado = True
    equipo.fecha_eliminacion = timezone.now()
    equipo.save()
    messages.success(request, f"Equipo '{equipo.serie}' movido a papelera.")
    return redirect('inventario')


@login_required
def ver_papelera(request):
    if not request.user.is_staff:
        return redirect('inventario')
    equipos_eliminados = Equipo.objects.filter(eliminado=True)
    return render(request, 'papelera.html', {'equipos': equipos_eliminados})


@login_required
def restaurar_equipo(request, pk):
    if not request.user.is_staff:
        return redirect('inventario')
    equipo = get_object_or_404(Equipo, pk=pk)
    equipo.eliminado = False
    equipo.save()
    return redirect('ver_papelera')


@login_required
def eliminar_permanente(request, pk):
    if not request.user.is_staff:
        return redirect('dashboard')
    equipo = get_object_or_404(Equipo, pk=pk)
    equipo.delete()
    return redirect('ver_papelera')


def exportar_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario Juana Cervantes"

    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('B2:F2')
    ws['B2'] = "I.E. JUANA CERVANTES DE BOLOGNESI"
    ws['B2'].font = Font(bold=True, size=16, color="003366")
    ws['B2'].alignment = center_align

    ws.merge_cells('B3:F3')
    ws['B3'] = f"REPORTE GENERAL DE INVENTARIO - Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['B3'].font = Font(italic=True, size=10)
    ws['B3'].alignment = center_align
    try:
        img = ExcelImage('static/inventario/logo.jpg')
        img.width = 60
        img.height = 60
        ws.add_image(img, 'A1')
    except:
        pass
    headers = ['SERIE', 'MARCA', 'MODELO', 'ESTADO', 'UBICACIÓN', 'RESPONSABLE']
    start_row = 5
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_style

    for row_num, e in enumerate(queryset, start_row + 1):
        row_data = [
            e.serie,
            e.marca,
            e.modelo,
            e.get_estado_display(),
            e.aula_actual if e.aula_actual else "No asignada",
            "---"
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border_style
            cell.alignment = Alignment(vertical="center", horizontal="left")
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 4
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Inventario_JC.xlsx'
    wb.save(response)
    return response


def exportar_pdf(queryset):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    p.setFillColor(colors.HexColor("#003366"))
    p.rect(0, h - 80, w, 80, fill=True, stroke=False)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, h - 45, "REPORTE DE INVENTARIO - I.E. JUANA CERVANTES")
    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 10)
    y = h - 120
    p.drawString(50, y, "Serie")
    p.drawString(180, y, "Modelo")
    p.drawString(380, y, "Estado")
    p.drawString(480, y, "Aula")
    p.line(50, y - 5, 550, y - 5)
    p.setFont("Helvetica", 9)
    y -= 25
    for e in queryset:
        p.drawString(50, y, str(e.serie))
        p.drawString(180, y, str(e.modelo)[:35])
        p.drawString(380, y, str(e.get_estado_display()))
        p.drawString(480, y, str(e.aula_actual if e.aula_actual else "N/A"))
        y -= 20
        if y < 50:
            p.showPage()
            y = h - 50
    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')


def exportar_libreta_pdf(request, estudiante_id, bimestre=None):
    estudiante = get_object_or_404(Estudiante, pk=estudiante_id)

    if request.user.is_authenticated:
        perfil = getattr(request.user, 'perfil', None)
        if not (request.user.is_staff or (perfil and perfil.rol == 'PADRE' and perfil.dni_hija == estudiante.dni)):
            messages.error(request, "No tienes permiso para ver estas notas.")
            return redirect('notas')
    else:
        messages.error(request, "Debes iniciar sesión.")
        return redirect('login')

    notas = Nota.objects.filter(estudiante=estudiante)
    if not notas.exists():
        messages.error(request, "No hay notas registradas para este estudiante.")
        return redirect('notas')
    bimestres = ['b1', 'b2', 'b3', 'b4']
    if bimestre and bimestre in bimestres:
        bimestres = [bimestre]
        titulo = f"Libreta de Calificaciones - {estudiante.apellidos}, {estudiante.nombres} - {bimestre.upper()}"
    else:
        titulo = f"Libreta de Calificaciones - {estudiante.apellidos}, {estudiante.nombres} - Todos los bimestres"
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 50
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, titulo)
    y -= 30
    p.setFont("Helvetica", 10)
    p.drawString(50, y, f"DNI: {estudiante.dni} | Grado: {estudiante.grado} | Sección: {estudiante.seccion}")
    y -= 30
    headers = ["Curso"] + [b.upper() for b in bimestres] + ["Promedio"]
    col_widths = [150] + [50] * len(bimestres) + [60]
    x_start = 50
    p.setFont("Helvetica-Bold", 9)
    x = x_start
    for i, header in enumerate(headers):
        p.drawString(x, y, header)
        x += col_widths[i]
    y -= 20
    p.line(x_start, y+5, x_start+sum(col_widths), y+5)
    p.setFont("Helvetica", 9)
    for nota in notas:
        x = x_start
        p.drawString(x, y, nota.curso[:25])
        x += col_widths[0]
        for b in bimestres:
            valor = getattr(nota, b, 0)
            p.drawString(x, y, f"{valor:.2f}")
            x += col_widths[1]
        if len(bimestres) == 1:
            promedio = getattr(nota, bimestres[0], 0)
        else:
            promedio = (nota.b1 + nota.b2 + nota.b3 + nota.b4) / 4
        p.drawString(x, y, f"{promedio:.2f}")
        y -= 20
        if y < 50:
            p.showPage()
            y = height - 50
            p.setFont("Helvetica-Bold", 9)
            x = x_start
            for i, header in enumerate(headers):
                p.drawString(x, y, header)
                x += col_widths[i]
            y -= 20
            p.line(x_start, y+5, x_start+sum(col_widths), y+5)
            p.setFont("Helvetica", 9)
    p.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    nombre_archivo = f"libreta_{estudiante.dni}_{bimestre if bimestre else 'todos'}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


def exportar_libreta_excel(request, estudiante_id, bimestre=None):
    estudiante = get_object_or_404(Estudiante, pk=estudiante_id)

    if request.user.is_authenticated:
        perfil = getattr(request.user, 'perfil', None)
        if not (request.user.is_staff or (perfil and perfil.rol == 'PADRE' and perfil.dni_hija == estudiante.dni)):
            messages.error(request, "No tienes permiso.")
            return redirect('notas')
    else:
        messages.error(request, "Debes iniciar sesión.")
        return redirect('login')
    notas = Nota.objects.filter(estudiante=estudiante)
    if not notas.exists():
        messages.error(request, "No hay notas.")
        return redirect('notas')
    wb = Workbook()
    ws = wb.active
    ws.title = "Libreta Notas"
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Libreta de Calificaciones - {estudiante.apellidos}, {estudiante.nombres}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"DNI: {estudiante.dni} | Grado: {estudiante.grado} | Sección: {estudiante.seccion}"
    ws['A2'].font = Font(italic=True)

    bimestres = ['b1', 'b2', 'b3', 'b4']
    if bimestre and bimestre in bimestres:
        bimestres = [bimestre]
    headers = ["Curso"] + [b.upper() for b in bimestres] + ["Promedio"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    row = 5
    for nota in notas:
        ws.cell(row=row, column=1, value=nota.curso)
        col = 2
        for b in bimestres:
            ws.cell(row=row, column=col, value=float(getattr(nota, b, 0)))
            col += 1
        if len(bimestres) == 1:
            promedio = float(getattr(nota, bimestres[0], 0))
        else:
            promedio = (nota.b1 + nota.b2 + nota.b3 + nota.b4) / 4
        ws.cell(row=row, column=col, value=round(promedio, 2))
        row += 1

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre_archivo = f"libreta_{estudiante.dni}_{bimestre if bimestre else 'todos'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response