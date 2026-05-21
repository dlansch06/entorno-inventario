import io
import json
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login as auth_login, authenticate, logout as auth_logout
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.db.models import Q
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from .models import Equipo, Estudiante, Actividad, Comunicado, Asistencia, Nota, Perfil
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


FAKE_NOTAS = [
    {
        'curso': 'Matemática',
        'b1': 15.0,
        'b2': 16.5,
        'b3': 14.5,
        'b4': 17.0,
        'promedio': 15.75
    },
    {
        'curso': 'Comunicación',
        'b1': 14.0,
        'b2': 15.0,
        'b3': 13.5,
        'b4': 16.0,
        'promedio': 14.625
    },
    {
        'curso': 'Historia',
        'b1': 17.0,
        'b2': 18.0,
        'b3': 16.5,
        'b4': 19.0,
        'promedio': 17.625
    },
]

FAKE_ASISTENCIAS = [
    {'fecha': timezone.now().date(), 'estado': 'Presente'},
    {'fecha': timezone.now().date() - timezone.timedelta(days=1), 'estado': 'Tardanza'},
    {'fecha': timezone.now().date() - timezone.timedelta(days=2), 'estado': 'Ausente'},
]

def login_view(request):
    if request.user.is_authenticated:
        if request.user.is_staff or request.user.is_superuser:
            return redirect('/admin/')
        return redirect('dashboard')

    if request.method == 'POST':
        action_type = request.POST.get('action_type', 'login_habitual')

        if action_type == 'login_habitual':
            form = AuthenticationForm(data=request.POST)
            if form.is_valid():
                user = form.get_user()
                auth_login(request, user)
                if user.is_staff or user.is_superuser:
                    return redirect('/admin/')
                if 'datos_reales' in request.session:
                    del request.session['datos_reales']
                return redirect('dashboard')
            else:
                messages.error(request, "Usuario o contraseña incorrectos.")

        elif action_type == 'registro_activacion':
            dni_padre = request.POST.get('dni_padre', '').strip()
            dni_hija = request.POST.get('dni_hija', '').strip()
            nueva_password = request.POST.get('nueva_password', '').strip()
            try:
                estudiante = Estudiante.objects.get(dni=dni_hija)
                user = User.objects.get(username=dni_padre)
                user.set_password(nueva_password)
                user.save()
                perfil, _ = Perfil.objects.get_or_create(user=user)
                perfil.rol = 'PADRE'
                perfil.dni = dni_padre
                perfil.dni_hija = dni_hija
                perfil.nivel_asignado = estudiante.nivel
                perfil.grado_asignado = estudiante.grado
                perfil.seccion_asignada = estudiante.seccion
                perfil.esta_aprobado = True
                perfil.save()
                estudiante.padre = user
                estudiante.save()
                user_autenticado = authenticate(request, username=dni_padre, password=nueva_password)
                if user_autenticado is not None:
                    auth_login(request, user_autenticado)
                    messages.success(request, f"¡Cuenta activada con éxito! Bienvenido al sistema.")
                    return redirect('notas')
            except Estudiante.DoesNotExist:
                messages.error(request, "El DNI de la estudiante no pertenece a ninguna alumna matriculada.")
            except User.DoesNotExist:
                messages.error(request, "El DNI del apoderado no figura registrado en el sistema.")
            except Exception as e:
                messages.error(request, f"Ocurrió un error en la sincronización: {str(e)}")
    return render(request, 'login.html')

def invitado_view(request):
    request.session['es_invitado'] = True
    return redirect('dashboard')

def logout_view(request):
    auth_logout(request)
    request.session.flush()
    return redirect('login')

def registro_padre_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            Perfil.objects.create(user=user)
            messages.success(request, "Cuenta creada. Espera la aprobación del Director para ver notas.")
            return redirect('dashboard')
    else:
        form = UserCreationForm()
    return render(request, 'padres/registro_padre.html', {'form': form})

@csrf_exempt
def procesar_acceso_padre_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            accion = data.get('accion')
            dni_padre = data.get('dniPadre')

            if accion == 'registrar':
                nombre_padre = data.get('nombrePadre')
                dni_hija = data.get('dniHija')

                try:
                    estudiante = Estudiante.objects.get(dni=dni_hija)
                except Estudiante.DoesNotExist:
                    return JsonResponse({
                        'status': 'ERROR_VALIDACION',
                        'message': 'El DNI de la estudiante no corresponde a ninguna alumna matriculada.'
                    }, status=404)

                user, created = User.objects.get_or_create(
                    username=dni_padre,
                    defaults={'first_name': nombre_padre}
                )
                if created:
                    user.set_password(dni_padre)
                    user.save()
                perfil, _ = Perfil.objects.get_or_create(user=user)
                perfil.rol = 'PADRE'
                perfil.dni = dni_padre
                perfil.dni_hija = dni_hija
                perfil.nivel_asignado = estudiante.nivel
                perfil.grado_asignado = estudiante.grado
                perfil.seccion_asignada = estudiante.seccion
                perfil.esta_aprobado = True
                perfil.fecha_aprobacion = timezone.now()
                perfil.save()
                estudiante.padre = user
                estudiante.save()
                auth_login(request, user)
                return JsonResponse({
                    'status': 'REGISTRO_EXITOSO',
                    'message': 'Cuenta vinculada, auto-llenada y verificada con éxito.',
                    'redirect_url': '/academico/notas/'
                })
        except Exception as e:
            return JsonResponse({'status': 'ERROR_INTERNO', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'METODO_NO_PERMITIDO'}, status=405)

def dashboard_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')
    return render(request, 'dashboard.html')

def acerca_de_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')
    return render(request, 'info.html')

def notas_view(request):
    notas_a_mostrar = FAKE_NOTAS.copy()
    estudiante = None
    error_msg = None
    es_invitado = not request.user.is_authenticated
    perfil = None
    son_datos_falsos = True
    promedios_bimestres = {}

    if request.user.is_authenticated:
        perfil = getattr(request.user, 'perfil', None)
        if not perfil:
            error_msg = "No existe perfil asociado."
        elif perfil.rol != 'PADRE':
            error_msg = "Este panel es exclusivo para padres."

    if request.method == 'POST':
        dni_manual = request.POST.get('documento')
        nombre_manual = request.POST.get('nombre')

        if dni_manual and nombre_manual:
            if es_invitado or not request.user.is_authenticated:
                error_msg = " Debes registrarte como padre para consultar calificaciones reales. Crea una cuenta e inicia sesión."
                notas_a_mostrar = FAKE_NOTAS
                estudiante = None
                son_datos_falsos = True
            else:
                try:
                    estudiante_encontrado = Estudiante.objects.get(dni=dni_manual)
                    nombre_completo = f"{estudiante_encontrado.nombres} {estudiante_encontrado.apellidos}".lower()
                    if nombre_manual.lower() in nombre_completo:
                        estudiante = estudiante_encontrado
                        notas_reales = Nota.objects.filter(estudiante=estudiante)
                        if notas_reales.exists():
                            notas_a_mostrar = [
                                {
                                    'curso': n.curso,
                                    'b1': float(n.b1),
                                    'b2': float(n.b2),
                                    'b3': float(n.b3),
                                    'b4': float(n.b4),
                                    'promedio': float(n.promedio)
                                }
                                for n in notas_reales
                            ]
                            son_datos_falsos = False
                            request.session['datos_reales'] = True
                        else:
                            notas_a_mostrar = []
                            son_datos_falsos = False
                            error_msg = "La estudiante no tiene notas registradas todavía."
                    else:
                        estudiante = None
                        notas_a_mostrar = FAKE_NOTAS
                        son_datos_falsos = True
                        error_msg = "El nombre no coincide con el número de DNI de la estudiante."
                except Estudiante.DoesNotExist:
                    estudiante = None
                    notas_a_mostrar = FAKE_NOTAS
                    son_datos_falsos = True
                    error_msg = "No se encontró ninguna estudiante con el DNI ingresado."
    if notas_a_mostrar and not son_datos_falsos:
        cantidad = len(notas_a_mostrar)
        if cantidad > 0:
            b1_sum = sum(n['b1'] for n in notas_a_mostrar)
            b2_sum = sum(n['b2'] for n in notas_a_mostrar)
            b3_sum = sum(n['b3'] for n in notas_a_mostrar)
            b4_sum = sum(n['b4'] for n in notas_a_mostrar)
            promedios_bimestres = {
                'b1': round(b1_sum / cantidad, 1),
                'b2': round(b2_sum / cantidad, 1),
                'b3': round(b3_sum / cantidad, 1),
                'b4': round(b4_sum / cantidad, 1),
            }
    elif notas_a_mostrar and son_datos_falsos:
        cantidad = len(notas_a_mostrar)
        if cantidad > 0:
            b1_sum = sum(n['b1'] for n in notas_a_mostrar)
            b2_sum = sum(n['b2'] for n in notas_a_mostrar)
            b3_sum = sum(n['b3'] for n in notas_a_mostrar)
            b4_sum = sum(n['b4'] for n in notas_a_mostrar)
            promedios_bimestres = {
                'b1': round(b1_sum / cantidad, 1),
                'b2': round(b2_sum / cantidad, 1),
                'b3': round(b3_sum / cantidad, 1),
                'b4': round(b4_sum / cantidad, 1),
            }
    promedio_base = 0
    if notas_a_mostrar:
        suma = sum(n['promedio'] for n in notas_a_mostrar)
        promedio_base = round(suma / len(notas_a_mostrar), 1)
    return render(
        request,
        'academico/notas.html',
        {
            'estudiante': estudiante,
            'perfil': perfil,
            'notas': notas_a_mostrar,
            'error': error_msg,
            'es_invitado': es_invitado,
            'promedio_base': promedio_base,
            'son_datos_falsos': son_datos_falsos,
            'promedios_bimestres': promedios_bimestres,
        }
    )

def asistencias_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')
    es_invitado = not request.user.is_authenticated
    asistencias_a_mostrar = FAKE_ASISTENCIAS.copy()
    estudiante = None
    error_msg = None
    son_datos_falsos = True

    if request.method == 'POST':
        dni_manual = request.POST.get('documento')
        nombre_manual = request.POST.get('nombre')
        if dni_manual and nombre_manual:
            if es_invitado or not request.user.is_authenticated:
                error_msg = " Debes registrarte como padre para consultar asistencias reales."
                asistencias_a_mostrar = FAKE_ASISTENCIAS
                son_datos_falsos = True
            else:
                try:
                    estudiante_encontrado = Estudiante.objects.get(dni=dni_manual)
                    nombre_completo = f"{estudiante_encontrado.nombres} {estudiante_encontrado.apellidos}".lower()
                    if nombre_manual.lower() in nombre_completo:
                        estudiante = estudiante_encontrado
                        asistencias_reales = Asistencia.objects.filter(estudiante=estudiante).order_by('-fecha')
                        if asistencias_reales.exists():
                            asistencias_a_mostrar = [
                                {'fecha': a.fecha.strftime('%d/%m/%Y'), 'estado': a.get_estado_display()}
                                for a in asistencias_reales
                            ]
                            son_datos_falsos = False
                        else:
                            asistencias_a_mostrar = []
                            son_datos_falsos = False
                            error_msg = "No hay registros de asistencia para esta estudiante."
                    else:
                        asistencias_a_mostrar = FAKE_ASISTENCIAS
                        son_datos_falsos = True
                        error_msg = "El nombre no coincide con el DNI."
                except Estudiante.DoesNotExist:
                    asistencias_a_mostrar = FAKE_ASISTENCIAS
                    son_datos_falsos = True
                    error_msg = "DNI no encontrado."
    return render(request, 'academico/asistencias.html', {
        'asistencias': asistencias_a_mostrar,
        'es_invitado': es_invitado,
        'error': error_msg,
        'son_datos_falsos': son_datos_falsos,
        'estudiante': estudiante,
    })

def comunicados_view(request):
    es_invitado = request.session.get('es_invitado', False)
    esta_autenticado = request.user.is_authenticated

    if not esta_autenticado and not es_invitado:
        return redirect('login')

    if esta_autenticado and hasattr(request.user, 'perfil') and request.user.perfil.rol == 'PADRE' and request.user.perfil.esta_aprobado:
        comunicados = Comunicado.objects.all().order_by('-fecha_publicacion')
        user_aprobado = True
    elif esta_autenticado:
        comunicados = Comunicado.objects.all().order_by('-fecha_publicacion')
        user_aprobado = True
    else:
        comunicados = Comunicado.objects.filter(es_general=True).order_by('-fecha_publicacion')
        user_aprobado = False

    return render(request, 'academico/comunicados.html', {
        'comunicados': comunicados,
        'user_aprobado': user_aprobado,
        'es_invitado': es_invitado
    })

def actividades_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')
    actividades = Actividad.objects.all().order_by('fecha_actividad')
    return render(request, 'academico/actividades.html', {'actividades': actividades})


def panel_invitado_calendario(request):
    actividades = Actividad.objects.filter(
        fecha_actividad__gte=timezone.now()
    ).order_by('fecha_actividad')
    return render(request, 'academico/actividades.html', {'actividades': actividades})

def inventario_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')

    equipos = Equipo.objects.filter(eliminado=False)
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')

    if q:
        equipos = equipos.filter(
            Q(serie__icontains=q) |
            Q(marca__icontains=q) |
            Q(modelo__icontains=q)
        )
    if estado:
        equipos = equipos.filter(estado=estado)

    return render(request, 'inventario.html', {
        'equipos': equipos,
        'q': q,
        'estado_actual': estado,
    })

def reportes_view(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')

    equipos = Equipo.objects.filter(eliminado=False)
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')

    if q:
        equipos = equipos.filter(
            Q(serie__icontains=q) |
            Q(marca__icontains=q) |
            Q(modelo__icontains=q)
        )
    if estado:
        equipos = equipos.filter(estado=estado)

    export_type = request.GET.get('export')
    if export_type == 'excel':
        return exportar_excel(equipos)
    elif export_type == 'pdf':
        return exportar_pdf(equipos)

    return render(request, 'reportes.html', {
        'equipos': equipos,
        'q': q,
        'estado_actual': estado
    })

@login_required
def mover_a_papelera(request, pk):
    if not request.user.is_staff:
        messages.error(request, "No tienes permiso para eliminar equipos.")
        return redirect('inventario')
    equipo = get_object_or_404(Equipo, pk=pk)
    equipo.eliminado = True
    equipo.fecha_eliminacion = timezone.now()
    equipo.save()
    messages.success(request, f"Equipo '{equipo.serie}' movido a papelera.")
    return redirect('inventario')

@login_required
def ver_papelera(request):
    if not request.user.is_staff:
        return redirect('inventario')
    equipos_eliminados = Equipo.objects.filter(eliminado=True)
    return render(request, 'papelera.html', {'equipos': equipos_eliminados})

@login_required
def restaurar_equipo(request, pk):
    if not request.user.is_staff:
        return redirect('inventario')
    equipo = get_object_or_404(Equipo, pk=pk)
    equipo.eliminado = False
    equipo.save()
    return redirect('ver_papelera')

@login_required
def eliminar_permanente(request, pk):
    if not request.user.is_staff:
        return redirect('dashboard')
    equipo = get_object_or_404(Equipo, pk=pk)
    equipo.delete()
    return redirect('ver_papelera')

def exportar_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario Juana Cervantes"

    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('B2:F2')
    ws['B2'] = "I.E. JUANA CERVANTES DE BOLOGNESI"
    ws['B2'].font = Font(bold=True, size=16, color="003366")
    ws['B2'].alignment = center_align

    ws.merge_cells('B3:F3')
    ws['B3'] = f"REPORTE GENERAL DE INVENTARIO - Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['B3'].font = Font(italic=True, size=10)
    ws['B3'].alignment = center_align
    try:
        img = ExcelImage('static/inventario/logo.jpg')
        img.width = 60
        img.height = 60
        ws.add_image(img, 'A1')
    except:
        pass
    headers = ['SERIE', 'MARCA', 'MODELO', 'ESTADO', 'UBICACIÓN', 'RESPONSABLE']
    start_row = 5
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_style

    for row_num, e in enumerate(queryset, start_row + 1):
        row_data = [
            e.serie,
            e.marca,
            e.modelo,
            e.get_estado_display(),
            e.aula_actual if e.aula_actual else "No asignada",
            "---"
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border_style
            cell.alignment = Alignment(vertical="center", horizontal="left")
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 4
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Inventario_JC.xlsx'
    wb.save(response)
    return response

def exportar_pdf(queryset):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4
    p.setFillColor(colors.HexColor("#003366"))
    p.rect(0, h - 80, w, 80, fill=True, stroke=False)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, h - 45, "REPORTE DE INVENTARIO - I.E. JUANA CERVANTES")
    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 10)
    y = h - 120
    p.drawString(50, y, "Serie")
    p.drawString(180, y, "Modelo")
    p.drawString(380, y, "Estado")
    p.drawString(480, y, "Aula")
    p.line(50, y - 5, 550, y - 5)
    p.setFont("Helvetica", 9)
    y -= 25
    for e in queryset:
        p.drawString(50, y, str(e.serie))
        p.drawString(180, y, str(e.modelo)[:35])
        p.drawString(380, y, str(e.get_estado_display()))
        p.drawString(480, y, str(e.aula_actual if e.aula_actual else "N/A"))
        y -= 20
        if y < 50:
            p.showPage()
            y = h - 50
    p.showPage()
    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

def nivel_academico(request):
    if not request.user.is_authenticated and not request.session.get('es_invitado'):
        return redirect('login')
    return render(request, 'academico/nivel_academico.html', {'nivel': 'SEC', 'nombre_nivel': 'Secundaria'})