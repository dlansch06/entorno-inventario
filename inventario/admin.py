import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
import os
from django.conf import settings
from django.http import HttpResponse
from django.contrib import admin
from django.utils.html import format_html # Para poner colores
from django.utils import timezone
from .models import Perfil, Equipo, Designacion, Estudiante, Nota, Asistencia, Comunicado, Actividad

@admin.action(description='Descargar Reporte Excel (Oficial)')
def exportar_excel_pro(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Estilos
    header_font = Font(name='Arial', bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4B2C20", end_color="4B2C20", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))

    
    logo_path = os.path.join(settings.BASE_DIR, 'static/img/logo.jpg') # Asegura esta ruta
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width, img.height = 80, 80
        ws.add_image(img, 'A1') 

    ws.merge_cells('B2:F2')
    ws['B2'] = "I.E. JUANA CERVANTES - REPORTE DE INVENTARIO"
    ws['B2'].font = Font(size=14, bold=True)
    ws['B2'].alignment = center_align

    headers = ['SERIE', 'CATEGORÍA', 'MARCA', 'MODELO', 'ESTADO', 'INGRESO COLEGIO']
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for row_num, obj in enumerate(queryset, 7):
        data = [
            obj.serie, obj.get_categoria_display(), obj.marca, obj.modelo,
            obj.estado, obj.fecha_ingreso_colegio.strftime('%d/%m/%Y') if obj.fecha_ingreso_colegio else ""
        ]
        for col_num, cell_value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.border = border

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.xlsx"'
    wb.save(response)
    return response

@admin.register(Equipo)
class EquipoAdmin(admin.ModelAdmin):
    list_display = ('serie', 'marca', 'modelo', 'color_estado', 'aula_actual', 'fecha_ingreso_colegio')
    list_filter = ('estado', 'categoria', 'fecha_ingreso_colegio')
    search_fields = ('serie', 'marca', 'modelo')
    actions = [exportar_excel_pro]

    def color_estado(self, obj):
        colores = {
            'DISPONIBLE': 'green',
            'EN_USO': 'blue',
            'MANTENIMIENTO': 'orange',
            'BAJA': 'red',
        }
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            colores.get(obj.estado, 'black'),
            obj.get_estado_display()
        )
    color_estado.short_description = 'Estado Actual'


@admin.action(description='Aprobar usuarios seleccionados')
def aprobar_usuarios(modeladmin, request, queryset):
    queryset.update(esta_aprobado=True, fecha_aprobacion=timezone.now())

@admin.register(Perfil)
class PerfilAdmin(admin.ModelAdmin):
    list_display = ('user', 'rol', 'esta_aprobado', 'grado_asignado', 'seccion_asignada')
    list_filter = ('rol', 'esta_aprobado', 'grado_asignado')
    list_editable = ('esta_aprobado',)
    actions = [aprobar_usuarios]

@admin.register(Estudiante)
class EstudianteAdmin(admin.ModelAdmin):
    list_display = ('apellidos', 'nombres', 'dni', 'nivel', 'grado', 'seccion', 'padre')
    list_filter = ('nivel', 'grado', 'seccion')
    search_fields = ('dni', 'apellidos')
    ordering = ('nivel', 'grado', 'seccion', 'apellidos')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser or request.user.perfil.rol in ['DIRECTOR', 'SECRETARIA', 'ADMIN']:
            return qs
        if request.user.perfil.rol == 'DOCENTE':
            return qs.filter(nivel=request.user.perfil.nivel_asignado,
                             grado=request.user.perfil.grado_asignado,
                             seccion=request.user.perfil.seccion_asignada)
        return qs.none()

@admin.register(Designacion)
class DesignacionAdmin(admin.ModelAdmin):
    list_display = ('equipo', 'persona_responsable', 'fecha_entrega', 'fecha_devolucion_real')
    list_filter = ('fecha_entrega', 'fecha_devolucion_real')

@admin.register(Nota)
class NotaAdmin(admin.ModelAdmin):
    list_display = ('estudiante', 'curso', 'b1', 'b2', 'b3', 'b4')
    list_editable =('b1', 'b2', 'b3', 'b4')
    list_filter = ('curso', 'estudiante__nivel', 'estudiante__grado', 'estudiante__seccion')
    def get_queryset(self, request):
        qs =super().get_queryset(request)
        if request.user.is_superuser or request.user.perfil.rol == ['ADMIN', 'SECRETARIA']:
            return qs
        if request.user.perfil.rol == 'DOCENTE':
            return qs.filter(estudiante__nivel=request.user.perfil.nivel_asignado,
                             estudiante__grado=request.user.perfil.grado_asignado,
                             estudiante__seccion=request.user.perfil.seccion_asignada)
        return qs.none()

@admin.register(Asistencia)
class AsistenciaAdmin(admin.ModelAdmin):
    list_display = ('estudiante', 'fecha', 'estado')
    list_filter = ('fecha', 'estado', 'estudiante__grado')
    list_editable = ('estado',)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser or request.user.perfil.rol in ['DIRECTOR', 'SECRETARIA', 'ADMIN']:
            return qs
        if request.user.perfil.rol == 'DOCENTE':
            return qs.filter(estudiante__nivel=request.user.perfil.nivel_asignado,
                             estudiante__grado=request.user.perfil.grado_asignado,
                             estudiante__seccion=request.user.perfil.seccion_asignada)
        return qs.none()
    
@admin.register(Comunicado)
class ComunicadoAdmin(admin.ModelAdmin):
    list_display = ('titulo', 'autor', 'es_general', 'fecha_publicacion')
    list_filter = ('es_general', 'fecha_publicacion')
    search_fields = ('titulo', 'contenido')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        user = request.user
        if user.is_superuser or user.perfil.rol in ['DIRECTOR', 'SECRETARIA', 'ADMIN']:
            return qs
        if user.perfil.rol == 'DOCENTE':
            return qs.filter(autor=user, es_general=False)
        return qs.none()

    def save_model(self, request, obj, form, change):
        if request.user.perfil.rol == 'DOCENTE':
            obj.es_general = False
            obj.autor = request.user
        elif not change and request.user.perfil.rol == 'DIRECTOR':
            obj.autor = request.user
        super().save_model(request, obj, form, change)

    def has_change_permission(self, request, obj=None):
        if obj and request.user.perfil.rol == 'SECRETARIA' and not obj.es_general:
            return False
        return super().has_change_permission(request, obj)
    
@admin.register(Actividad)
class ActividadAdmin(admin.ModelAdmin):
    list_display = ('titulo', 'tipo', 'fecha_actividad')
    list_filter = ('tipo', 'fecha_actividad')
    search_fields = ('titulo', 'descripcion')

    def has_module_permission(self, request):
        # Solo Director, Secretaria y Admin manejan Actividades
        if request.user.is_superuser or request.user.perfil.rol in ['DIRECTOR', 'SECRETARIA', 'ADMIN']:
            return True
        return False
    